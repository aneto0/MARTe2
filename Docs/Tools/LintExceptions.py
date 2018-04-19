#!/usr/local/bin/python2.7

import os
import re
import argparse
from openpyxl import Workbook

parser = argparse.ArgumentParser(description='Collects linting exceptions documented in source code.')
parser.add_argument(dest='rootDir', nargs=1, help='The root of the source code.')

rootDir = parser.parse_args().rootDir[0]

print(rootDir)

# regex patterns

# Regex for C/C++ style comments
# It works also with multiline comments
c_comments = re.compile(
  r'//.*?$|/\*.*?\*/|\'(?:\\.|[^\\\'])*\'|"(?:\\.|[^\\"])*"',
  re.DOTALL | re.MULTILINE
)

lint_command = re.compile(r'(-e.*?[\({].*?[\)}])')
 
# Set the directory you want to start from
lint_exceptions = dict()
for dirName, subdirList, fileList in os.walk(rootDir):
    print('Looking in: %s' % dirName)
    for fname in fileList:
        # Look only for C/C++ h or cpp files
        if (fname[len(fname)-2:] == '.h') or (fname[len(fname)-4:] == '.cpp'):
            f = file(os.path.join(dirName,fname)).read()

            # Find all comments in the file
            occurrences = re.findall(c_comments,f)

            # For each occurrence filter the ones starting with /*lint
            for o in occurrences:
                if '/*lint' in o:
                    # Create a key in the dict for the file if it does not exists
                    if fname not in lint_exceptions:
                        lint_exceptions[fname] = dict()

                    # Split the comment in an array of lines (needed for multiline comments)
                    # and get first line
                    first_comment_line = o.splitlines()[0]
                    # Search the first line to find line number
                    for n,line in enumerate(open(os.path.join(dirName,fname))):
                        if first_comment_line in line:
                            linenum = n+1

                    lint_exceptions[fname][linenum] = dict()
                    lint_exceptions[fname][linenum]['commands'] = []
                    lint_exceptions[fname][linenum]['comment'] = o

                    list_of_commands = re.findall(lint_command,o)
                    if list_of_commands is not None:
                        for command in list_of_commands:
                            lint_exceptions[fname][linenum]['commands'].append(command)

#print(lint_exceptions)

# Create XLS
wb = Workbook()
ws = wb.active
ws['A1'] = 'File'
ws['B1'] = 'Line'
ws['C1'] = 'Exception'
ws['D1'] = 'Comment'

current_line = 2
for fname,lines in lint_exceptions.iteritems():
    cell = 'A' + str(current_line)
    ws[cell] = fname

    for lineno,data in lines.iteritems():
        cell = 'B' + str(current_line)
        ws[cell] = str(lineno)

        cell = 'D' + str(current_line)
        ws[cell] = data['comment']

        for command in data['commands']:
            cell = 'C' + str(current_line)
            ws[cell] = command
            current_line = current_line+1

wb.save("lint.xlsx")

